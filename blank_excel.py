import datetime
import calendar
from openpyxl.styles import PatternFill, Font
import pandas as pd
from doctor_data import BLANK_DOCTOR_LIST, THAI_HOLIDAYS

def generate_blank_excel(year, month, filename="blank_schedule.xlsx"):
    days_in_month = calendar.monthrange(year, month)[1]
    dates = [datetime.date(year, month, d) for d in range(1, days_in_month+1)]
    data = []
    num_doctors = len(BLANK_DOCTOR_LIST)
    for idx, date in enumerate(dates):
        day_of_week = date.strftime("%a")
        er_doctor = BLANK_DOCTOR_LIST[idx % num_doctors] if num_doctors > 0 else ""
        is_weekend = date.weekday() >= 5
        is_holiday = date in THAI_HOLIDAYS
        has_opd = not is_weekend and not is_holiday
        data.append({
            "Date": date.strftime("%Y-%m-%d"),
            "Day of week": day_of_week,
            "ER": "",
            "OPD1": "",
            "OPD2": "",
            "OPD3/LR/ANC": "",
            "OPD4": "ประภาส" if has_opd else "",
            "เรื้อรัง": "สมชาย" if has_opd else "",
            "Wardหญิง": "",
            "Wardชาย": "",
            "เวร ER": er_doctor,
            "เวร Ward": ""
        })
    df = pd.DataFrame(data)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Schedule")
        ws = writer.sheets["Schedule"]
        # Insert blank column after main data
        ws.insert_cols(df.shape[1]+1)
        ws.cell(row=1, column=df.shape[1]+1, value="")

        # Color header row (column headers) with blue and clear border
        header_blue = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
        for col in range(1, df.shape[1]+1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_blue
            cell.border = None

        # Color weekend and holiday rows using date column, and clear border
        light_orange = PatternFill(start_color="FFF4B084", end_color="FFF4B084", fill_type="solid")
        light_red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        for idx in range(2, len(data)+2):
            date_str = ws.cell(row=idx, column=1).value
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            is_holiday = date_obj in THAI_HOLIDAYS
            is_wkend = date_obj.weekday() >= 5 or is_holiday
            if is_holiday:
                for col in range(1, df.shape[1]+1):
                    cell = ws.cell(row=idx, column=col)
                    cell.fill = light_red
                    cell.border = None
            elif is_wkend:
                for col in range(1, df.shape[1]+1):
                    cell = ws.cell(row=idx, column=col)
                    cell.fill = light_orange
                    cell.border = None

        # New count table header (2 rows: merged weekday/weekend, then เวร ER/เวร Ward)
        start_row = 1
        start_col = df.shape[1]+3

        # First header row: Doctor | Weekday (merged 3 cols) | Weekend (merged 3 cols) | Total (merged 1 col)
        bold_font = Font(bold=True)
        ws.cell(row=start_row, column=start_col, value="Doctor").font = bold_font
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row+1, end_column=start_col)
        ws.cell(row=start_row, column=start_col+1, value="Weekday").font = bold_font
        ws.merge_cells(start_row=start_row, start_column=start_col+1, end_row=start_row, end_column=start_col+3)
        ws.cell(row=start_row, column=start_col+4, value="Weekend").font = bold_font
        ws.merge_cells(start_row=start_row, start_column=start_col+4, end_row=start_row, end_column=start_col+6)
        ws.cell(row=start_row, column=start_col+7, value="Total").font = bold_font
        ws.merge_cells(start_row=start_row, start_column=start_col+7, end_row=start_row+1, end_column=start_col+7)

        # Second header row: เวร ER | เวร Ward | Total (Weekday) | เวร ER | เวร Ward | Total (Weekend)
        ws.cell(row=start_row+1, column=start_col+1, value="เวร ER")
        ws.cell(row=start_row+1, column=start_col+2, value="เวร Ward")
        ws.cell(row=start_row+1, column=start_col+3, value="Total")
        ws.cell(row=start_row+1, column=start_col+4, value="เวร ER")
        ws.cell(row=start_row+1, column=start_col+5, value="เวร Ward")
        ws.cell(row=start_row+1, column=start_col+6, value="Total")

        # Precompute cell references for each (col_name, is_weekend) combination
        col_types = [
            ("เวร ER", False, 1),
            ("เวร Ward", False, 2),
            ("เวร ER", True, 4),
            ("เวร Ward", True, 5)
        ]
        cell_refs = { (col_name, is_weekend): [] for (col_name, is_weekend, _) in col_types }
        for row_idx in range(2, len(data)+2):
            date_str = ws.cell(row=row_idx, column=1).value
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            is_holiday = date_obj in THAI_HOLIDAYS
            is_wkend = (date_obj.weekday() >= 5 or is_holiday)
            for col_name, is_weekend, _ in col_types:
                if is_weekend == is_wkend:
                    col_letter = chr(ord('A') + list(data[0].keys()).index(col_name))
                    cell_ref = f'{col_letter}{row_idx}'
                    cell_refs[(col_name, is_weekend)].append(cell_ref)

        # For each doctor, count เวร ER/Ward for weekday/weekend, and add totals (ER before Ward)
        for d_idx, doctor in enumerate(BLANK_DOCTOR_LIST):
            row_num = start_row+2+d_idx
            ws.cell(row=row_num, column=start_col, value=doctor)
            # Weekday counts
            weekday_er_cells = cell_refs[("เวร ER", False)]
            weekday_ward_cells = cell_refs[("เวร Ward", False)]
            weekend_er_cells = cell_refs[("เวร ER", True)]
            weekend_ward_cells = cell_refs[("เวร Ward", True)]
            doctor_cell = ws.cell(row=row_num, column=start_col).coordinate
            # Weekday ER
            if weekday_er_cells:
                ws.cell(row=row_num, column=start_col+1, value=f'=SUM({', '.join([f'--({cell}={doctor_cell})' for cell in weekday_er_cells])})')
            else:
                ws.cell(row=row_num, column=start_col+1, value='=0')
            # Weekday Ward
            if weekday_ward_cells:
                ws.cell(row=row_num, column=start_col+2, value=f'=SUM({', '.join([f'--({cell}={doctor_cell})' for cell in weekday_ward_cells])})')
            else:
                ws.cell(row=row_num, column=start_col+2, value='=0')
            # Weekday Total
            ws.cell(row=row_num, column=start_col+3, value=f'=SUM({ws.cell(row=row_num, column=start_col+1).coordinate},{ws.cell(row=row_num, column=start_col+2).coordinate})')
            # Weekend ER
            if weekend_er_cells:
                ws.cell(row=row_num, column=start_col+4, value=f'=SUM({', '.join([f'--({cell}={doctor_cell})' for cell in weekend_er_cells])})')
            else:
                ws.cell(row=row_num, column=start_col+4, value='=0')
            # Weekend Ward
            if weekend_ward_cells:
                ws.cell(row=row_num, column=start_col+5, value=f'=SUM({', '.join([f'--({cell}={doctor_cell})' for cell in weekend_ward_cells])})')
            else:
                ws.cell(row=row_num, column=start_col+5, value='=0')
            # Weekend Total
            ws.cell(row=row_num, column=start_col+6, value=f'=SUM({ws.cell(row=row_num, column=start_col+4).coordinate},{ws.cell(row=row_num, column=start_col+5).coordinate})')
            # Grand Total
            ws.cell(row=row_num, column=start_col+7, value=f'=SUM({ws.cell(row=row_num, column=start_col+3).coordinate},{ws.cell(row=row_num, column=start_col+6).coordinate})')
    print(f"Blank schedule saved to {filename}")
