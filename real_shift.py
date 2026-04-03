import argparse
import datetime
import os
from collections import defaultdict
from ortools.sat.python import cp_model
from constraints import is_weekend, is_holiday
from doctor_data import THAI_HOLIDAYS

# ─────────────────────────────────────────────────────────
#  Doctor roster and month configuration
# ─────────────────────────────────────────────────────────

doctor_data = {
    "ธนัท": {
        "weekday": {"ER": 3, "ward": 3},
        "weekend": {"ER": 2, "ward": 2},
    },
    "กุลประวีณ์": {
        "weekday": {"ER": 3, "ward": 3},
        "weekend": {"ER": 2, "ward": 2},
    },
    "สุประวีณ์": {
        "weekday": {"ER": 3, "ward": 3},
        "weekend": {"ER": 2, "ward": 2},
    },
    "กุลพักตร์": {
        "weekday": {"ER": 3, "ward": 3},
        "weekend": {"ER": 2, "ward": 2},
    },
    "พัชรพร": {
        "weekday": {"ER": 3, "ward": 3},
        "weekend": {"ER": 2, "ward": 2},
    },
    "ฤชุกร": {
        "weekday": {"ER": 3, "ward": 3},
        "weekend": {"ER": 2, "ward": 2},
    },
}

Month = 4
Year = 2026

# On these dates the same doctor covers BOTH ER and ward
date_doubles = [
    datetime.date(2026, 4, 3),
    datetime.date(2026, 4, 10),
    datetime.date(2026, 4, 17),
    datetime.date(2026, 4, 24),
]

doctor_date_off = {
    "ธนัท": [
        # 22-26 April
        datetime.date(2026, 4, 22),
        datetime.date(2026, 4, 23),
        datetime.date(2026, 4, 24),
        datetime.date(2026, 4, 25),
        datetime.date(2026, 4, 26),
    ],
    "กุลประวีณ์": [
        # 4-13, 29-30 April
        datetime.date(2026, 4, 4),
        datetime.date(2026, 4, 5),
        datetime.date(2026, 4, 6),
        datetime.date(2026, 4, 7),
        datetime.date(2026, 4, 8),
        datetime.date(2026, 4, 9),
        datetime.date(2026, 4, 10),
        datetime.date(2026, 4, 11),
        datetime.date(2026, 4, 12),
        datetime.date(2026, 4, 13),
        datetime.date(2026, 4, 28),
        datetime.date(2026, 4, 29),
        datetime.date(2026, 4, 30),
    ],
    "สุประวีณ์": [
        # 4-5, 16-24 April
        datetime.date(2026, 4, 3),
        datetime.date(2026, 4, 4),
        datetime.date(2026, 4, 5),
        datetime.date(2026, 4, 14),
        datetime.date(2026, 4, 15),
        datetime.date(2026, 4, 16),
        datetime.date(2026, 4, 17),
        datetime.date(2026, 4, 18),
        datetime.date(2026, 4, 19),
        datetime.date(2026, 4, 20),
        datetime.date(2026, 4, 21),
        datetime.date(2026, 4, 22),
        datetime.date(2026, 4, 23),
        datetime.date(2026, 4, 24),
        datetime.date(2026, 4, 25),
    ],
    "กุลพักตร์": [
        # 4-5, 16-22 April
        datetime.date(2026, 4, 3),
        datetime.date(2026, 4, 4),
        datetime.date(2026, 4, 5),
        datetime.date(2026, 4, 14),
        datetime.date(2026, 4, 15),
        datetime.date(2026, 4, 16),
        datetime.date(2026, 4, 17),
        datetime.date(2026, 4, 18),
        datetime.date(2026, 4, 19),
        datetime.date(2026, 4, 20),
        datetime.date(2026, 4, 21),
        datetime.date(2026, 4, 22),
    ],
    "พัชรพร": [
        # 1-6, 17-19, 29-30 April
        datetime.date(2026, 4, 1),
        datetime.date(2026, 4, 2),
        datetime.date(2026, 4, 3),
        datetime.date(2026, 4, 4),
        datetime.date(2026, 4, 5),
        datetime.date(2026, 4, 6),
        datetime.date(2026, 4, 17),
        datetime.date(2026, 4, 18),
        datetime.date(2026, 4, 19),
        datetime.date(2026, 4, 28),
        datetime.date(2026, 4, 29),
        datetime.date(2026, 4, 30),
    ],
    "ฤชุกร": [
        # 8-15 April
        datetime.date(2026, 4, 7),
        datetime.date(2026, 4, 8),
        datetime.date(2026, 4, 9),
        datetime.date(2026, 4, 10),
        datetime.date(2026, 4, 11),
        datetime.date(2026, 4, 12),
        datetime.date(2026, 4, 13),
        datetime.date(2026, 4, 14),
        datetime.date(2026, 4, 15),
    ]
}


# ─────────────────────────────────────────────────────────
#  Schedule generator
# ─────────────────────────────────────────────────────────

def generate_real_schedule(year, month, doctor_data, date_doubles,
                           doctor_date_off=None, time_limit_seconds=60):
    """
    Generate a simple daily schedule with exactly one ER and one ward
    shift per day using OR-Tools CP-SAT.

    Constraints
    -----------
    1. Each day has exactly one doctor assigned to ER and one to ward.
    2. On date_doubles, the same doctor covers both ER and ward.
    3. No doctor works on two consecutive calendar days.

    Objective
    ---------
    Minimise the imbalance (max - min) in total shifts across doctors.
    """
    doctors = list(doctor_data.keys())
    n_doc = len(doctors)

    days_in_month = (
        datetime.date(year, month % 12 + 1, 1) - datetime.timedelta(days=1)
    ).day
    days = [datetime.date(year, month, d) for d in range(1, days_in_month + 1)]
    double_set = set(date_doubles)
    off_sets = {doc: set(dates) for doc, dates in (doctor_date_off or {}).items()}

    model = cp_model.CpModel()

    # Decision variables
    # er[d][i]   = 1 iff doctor i works ER on day d
    # ward[d][i] = 1 iff doctor i works ward on day d
    er   = {d: [model.NewBoolVar(f"er_d{d.day}_doc{i}")   for i in range(n_doc)] for d in days}
    ward = {d: [model.NewBoolVar(f"ward_d{d.day}_doc{i}") for i in range(n_doc)] for d in days}

    # ── Constraint 0: doctor cannot work on their day off ──────────────
    for i, doc in enumerate(doctors):
        for d in days:
            if d in off_sets.get(doc, set()):
                model.Add(er[d][i] == 0)
                model.Add(ward[d][i] == 0)

    # ── Constraint 0b: pinned assignments ─────────────────────────────
    pinned = {
        datetime.date(year, month, 10): "ธนัท",
        datetime.date(year, month, 15): "กุลประวีณ์",
        datetime.date(year, month, 22): "กุลประวีณ์",
    }
    for pin_date, pin_doc in pinned.items():
        if pin_date in er:
            pi = doctors.index(pin_doc)
            model.Add(er[pin_date][pi] + ward[pin_date][pi] >= 1)

    # ── Constraint 0c: shift-type restrictions per doctor/date ─────────
    # Each entry: (day, doctor, shift) where shift is "ER" or "ward"
    no_shift = [
        (1, "ฤชุกร", "ER"),
        (2, "ฤชุกร", "ER"),
        (5, "ธนัท", "ER"),
        (7, "ธนัท", "ER"),
        (8, "ธนัท", "ER"),
        (9, "ธนัท", "ER"),
        (28, "สุประวีณ์", "ER"),
        (29, "สุประวีณ์", "ER"),
        # (30, "สุประวีณ์", "ER"),
    ]
    for day_num, doc, shift_type in no_shift:
        d = datetime.date(year, month, day_num)
        if d in er:
            di = doctors.index(doc)
            if shift_type == "ER":
                model.Add(er[d][di] == 0)
            else:
                model.Add(ward[d][di] == 0)

    # ── Constraint 1: exactly one doctor per shift per day ────────────
    for d in days:
        model.AddExactlyOne(er[d])
        model.AddExactlyOne(ward[d])

    # ── Constraint 2: date_doubles → same doctor for ER and ward ──────
    for d in days:
        if d in double_set:
            for i in range(n_doc):
                model.Add(er[d][i] == ward[d][i])

    # ── Constraint 2b: weekend/holiday → different doctors for ER and ward ──
    for d in days:
        if d not in double_set and (is_weekend(d) or is_holiday(d)):
            for i in range(n_doc):
                model.Add(er[d][i] + ward[d][i] <= 1)

    # ── Constraint 3a: no more than 1 ER shift per doctor in any 3 consecutive days ──
    for k in range(len(days) - 2):
        d0, d1, d2 = days[k], days[k + 1], days[k + 2]
        for i in range(n_doc):
            model.Add(er[d0][i] + er[d1][i] + er[d2][i] <= 1)
    
    # ── Constraint 3b: no doctor on consecutive ward days ──────────────
    for k in range(len(days) - 1):
        d_cur  = days[k]
        d_next = days[k + 1]
        for i in range(n_doc):
            model.Add(ward[d_cur][i] + ward[d_next][i] <= 1)

    # ── Constraint 3c: no weekday shift immediately after a weekend/holiday ──
    for k in range(len(days) - 1):
        d_cur  = days[k]
        d_next = days[k + 1]
        if (is_weekend(d_cur) or is_holiday(d_cur)) and not (is_weekend(d_next) or is_holiday(d_next)):
            for i in range(n_doc):
                worked_cur  = model.NewBoolVar(f"worked_wkend_d{d_cur.day}_doc{i}")
                model.AddMaxEquality(worked_cur, [er[d_cur][i], ward[d_cur][i]])
                model.Add(worked_cur + er[d_next][i] + ward[d_next][i] <= 1)

    # ── Constraint 4: no doctor works more than 2 consecutive days ────
    for k in range(len(days) - 2):
        d0, d1, d2 = days[k], days[k + 1], days[k + 2]
        for i in range(n_doc):
            worked0 = model.NewBoolVar(f"worked3_d{d0.day}_doc{i}")
            worked1 = model.NewBoolVar(f"worked3_d{d1.day}_doc{i}")
            worked2 = model.NewBoolVar(f"worked3_d{d2.day}_doc{i}")
            model.AddMaxEquality(worked0, [er[d0][i], ward[d0][i]])
            model.AddMaxEquality(worked1, [er[d1][i], ward[d1][i]])
            model.AddMaxEquality(worked2, [er[d2][i], ward[d2][i]])
            model.Add(worked0 + worked1 + worked2 <= 2)

    # ── Constraint 4b: max 3 shifts per doctor during days 11-15 ──────
    window_days = [d for d in days if 11 <= d.day <= 15]
    for i in range(n_doc):
        model.Add(sum(er[d][i] + ward[d][i] for d in window_days) <= 3)

    # ── Constraint 4c: ธนัท must have exactly 2 shifts during days 11-15 ──
    idx_thanat = doctors.index("ธนัท")
    model.Add(sum(er[d][idx_thanat] + ward[d][idx_thanat] for d in window_days) == 2)

    # ── Constraint 5: enforce shift counts from doctor_data ───────────
    # Weekday days = not weekend and not holiday; weekend days = weekend or holiday
    weekday_days = [d for d in days if not is_weekend(d) and not is_holiday(d)]
    wkend_days   = [d for d in days if is_weekend(d) or is_holiday(d)]

    for i, doc in enumerate(doctors):
        quota = doctor_data[doc]
        # ER on weekdays
        model.Add(sum(er[d][i]   for d in weekday_days) == quota["weekday"]["ER"])
        # ward on weekdays
        model.Add(sum(ward[d][i] for d in weekday_days) == quota["weekday"]["ward"])
        # ER on weekends/holidays
        model.Add(sum(er[d][i]   for d in wkend_days)   == quota["weekend"]["ER"])
        # ward on weekends/holidays
        model.Add(sum(ward[d][i] for d in wkend_days)   == quota["weekend"]["ward"])

    # ── Objective: maximise co-work days for สุประวีณ์ + กุลพักตร์ ──────
    idx_su  = doctors.index("สุประวีณ์")
    idx_kul = doctors.index("กุลพักตร์")
    co_work_vars = []
    for d in days:
        if is_weekend(d) or is_holiday(d):
            su_works  = model.NewBoolVar(f"su_works_d{d.day}")
            kul_works = model.NewBoolVar(f"kul_works_d{d.day}")
            model.AddMaxEquality(su_works,  [er[d][idx_su],  ward[d][idx_su]])
            model.AddMaxEquality(kul_works, [er[d][idx_kul], ward[d][idx_kul]])
            both = model.NewBoolVar(f"both_work_d{d.day}")
            model.AddMinEquality(both, [su_works, kul_works])
            co_work_vars.append(both)

    co_work_count = model.NewIntVar(0, len(co_work_vars) if co_work_vars else 1, "co_work_count")
    if co_work_vars:
        model.Add(co_work_count == sum(co_work_vars))
    else:
        model.Add(co_work_count == 0)

    # ── Objective term: minimise max double-day assignments per doctor ─
    # On a double day the assigned doctor works both ER and ward (er==ward==1).
    double_days_list = [d for d in days if d in double_set]
    doc_double_counts = []
    for i in range(n_doc):
        # On a double day er[d][i] == ward[d][i], so er[d][i] alone == 1 iff assigned
        cnt = model.NewIntVar(0, len(double_days_list), f"double_cnt_doc{i}")
        model.Add(cnt == sum(er[d][i] for d in double_days_list))
        doc_double_counts.append(cnt)

    max_double = model.NewIntVar(0, len(double_days_list), "max_double")
    model.AddMaxEquality(max_double, doc_double_counts)

    # Combined objective:
    #   primary   : maximise co-work days         (weight 100)
    #   secondary : minimise max double-day count  (weight 1)
    model.Maximize(co_work_count - max_double)

    # ── Solve ─────────────────────────────────────────────────────────
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit_seconds
    solver.parameters.log_search_progress = False

    print("[OR-Tools] Solving real shift schedule...")
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        print(f"[OR-Tools] No solution found. Status: {solver.StatusName(status)}")
        return None

    print(f"[OR-Tools] Status    : {solver.StatusName(status)}")
    print(f"[OR-Tools] Wall time : {solver.WallTime():.2f}s")
    print(f"[OR-Tools] Co-work   : {solver.Value(co_work_count)} weekend/holiday days สุประวีณ์+กุลพักตร์")
    print(f"[OR-Tools] Max double: {solver.Value(max_double)} double-day shifts (max per doctor)")

    # ── Extract schedule ───────────────────────────────────────────────
    schedule    = {}
    shift_count = defaultdict(int)

    for d in days:
        er_doc   = next(doctors[i] for i in range(n_doc) if solver.Value(er[d][i]))
        ward_doc = next(doctors[i] for i in range(n_doc) if solver.Value(ward[d][i]))
        schedule[d] = {"ER": er_doc, "ward": ward_doc}
        shift_count[er_doc]   += 1
        shift_count[ward_doc] += 1

    return schedule, dict(shift_count)


# ─────────────────────────────────────────────────────────
#  Pretty-print helper
# ─────────────────────────────────────────────────────────

def print_schedule(schedule, shift_count):
    DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    print("\n" + "=" * 60)
    print(f"  Schedule  {Year}/{Month:02d}")
    print("=" * 60)
    print(f"  {'Date':<14} {'Day':<5} {'ER':<16} {'Ward':<16} {'Note'}")
    print("-" * 60)

    for date in sorted(schedule):
        entry    = schedule[date]
        day_name = DAY_NAMES[date.weekday()]

        notes = []
        if date in set(date_doubles):
            notes.append("double")
        if date in THAI_HOLIDAYS:
            notes.append("holiday")
        elif date.weekday() >= 5:
            notes.append("weekend")

        note_str = ", ".join(notes)
        print(f"  {str(date):<14} {day_name:<5} {entry['ER']:<16} {entry['ward']:<16} {note_str}")

    print("=" * 60)
    print("\nTotal shift slots per doctor:")
    for doc, cnt in sorted(shift_count.items(), key=lambda x: -x[1]):
        print(f"  {doc}: {cnt}")


# ─────────────────────────────────────────────────────────
#  Excel export
# ─────────────────────────────────────────────────────────

def save_real_schedule_to_xlsx(schedule, shift_count, filename="real_schedule.xlsx"):
    """Export the real-shift schedule to a styled Excel file."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    thin      = Side(style="thin")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    orange_fill = PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")
    green_fill  = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
    header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Header row
    headers = ["Date", "Day", "Type", "ER", "Ward"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, date in enumerate(sorted(schedule), 2):
        entry    = schedule[date]
        day_name = DAY_NAMES[date.weekday()]
        is_wkend = is_weekend(date) or is_holiday(date)
        fill     = orange_fill if is_wkend else green_fill

        notes = []
        if date in set(date_doubles):
            notes.append("double")
        if date in THAI_HOLIDAYS:
            notes.append("holiday")
        elif date.weekday() >= 5:
            notes.append("weekend")
        day_type = ", ".join(notes) if notes else "weekday"

        for col, val in enumerate([str(date), day_name, day_type, entry["ER"], entry["ward"]], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Summary: shift counts per doctor
    sc = 7
    for col, h in enumerate(["Doctor", "Total Shifts"], sc):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for i, (doc, cnt) in enumerate(sorted(shift_count.items(), key=lambda x: -x[1]), 2):
        ws.cell(row=i, column=sc,   value=doc)
        ws.cell(row=i, column=sc+1, value=cnt)

    # Column widths
    for col, width in zip(range(1, 9), [14, 6, 12, 16, 16, 2, 16, 14]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(filename)
    print(f"Schedule saved to {os.path.abspath(filename)}")


# ─────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate real shift schedule")
    parser.add_argument("-o", "--output", metavar="FILE",
                        help="Save schedule to Excel file (e.g. schedule.xlsx)")
    parser.add_argument("--time-limit", type=int, default=60,
                        metavar="SECONDS", help="Solver time limit in seconds (default: 60)")
    args = parser.parse_args()

    result = generate_real_schedule(Year, Month, doctor_data, date_doubles,
                                    doctor_date_off=doctor_date_off,
                                    time_limit_seconds=args.time_limit)
    if result:
        schedule, shift_count = result
        print_schedule(schedule, shift_count)
        if args.output:
            save_real_schedule_to_xlsx(schedule, shift_count, filename=args.output)

