import pandas as pd
import os
from doctor_data import DOCTOR_DATA, DOCTOR_AUTOPSY_DATA, SHIFT_TIMES, adjust_doctor_data
from constraints import is_weekend, is_holiday
from openpyxl.styles import PatternFill, Font, Border, Side


def transform_autopsy_data(autopsy_data):
    THAI_SHIFT_TIMES = {
        SHIFT_TIMES["NIGHT"]: "ด",
        SHIFT_TIMES["DAY"]: "ช",
        SHIFT_TIMES["EVENING"]: "บ",
    }
    transformed = {}
    for doctor, dates in autopsy_data.items():
        for date, shift_time in dates:
            if date not in transformed:
                transformed[date] = []
            transformed[date].append(
                f"{THAI_SHIFT_TIMES[shift_time]}/{doctor}")
    return transformed


def save_schedule_to_xlsx(schedule, filename="schedule.xlsx"):
    all_shifts = set()
    for date in schedule:
        for shift_type, shift_time, _ in schedule[date]:
            all_shifts.add((shift_type, shift_time))
    all_shifts = sorted(all_shifts, key=lambda x: (x[0], x[1]))
    data = []
    dates = sorted(schedule.keys())
    for date in dates:
        row = {f"{stype} {stime}": "" for stype, stime in all_shifts}
        for shift_type, shift_time, doctor in schedule[date]:
            row[f"{shift_type} {shift_time}"] = doctor
        row = {
            "Period": "Weekend" if is_weekend(date) or is_holiday(date) else "Weekday",
            "Day": date.strftime("%A")
        } | row
        data.append(row)
    df = pd.DataFrame(data, index=dates)
    df.index.name = "Date"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Schedule")
        ws = writer.sheets["Schedule"]
        # write column for autopsy data
        if DOCTOR_AUTOPSY_DATA:
            bold_font = Font(bold=True)
            autopsy_data = transform_autopsy_data(DOCTOR_AUTOPSY_DATA)
            autopsy_column = len(df.columns) + 2
            header = ws.cell(row=1, column=autopsy_column, value="Autopsy")
            header.font = bold_font
            header.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                   top=Side(style="thin"), bottom=Side(style="thin"))
            for i, date in enumerate(dates):
                autopsy_info = autopsy_data.get(date, [])
                ws.cell(row=i + 2, column=autopsy_column,
                        value=", ".join(d for d in autopsy_info))
        # write expected shifts
        start_col = len(df.columns) + 4
        for row in range(1, len(df) + 2):
            ws.cell(row=row, column=start_col-1, value=None)
        adjusted_doctor_data = adjust_doctor_data(DOCTOR_DATA)
        doctors = list(adjusted_doctor_data.keys())
        ws.cell(row=1, column=start_col, value="Doctor")
        ws.cell(row=1, column=start_col+1, value="Weekday ER")
        ws.cell(row=1, column=start_col+2, value="Weekday ward")
        ws.cell(row=1, column=start_col+3, value="Total Weekday")
        ws.cell(row=1, column=start_col+4, value="Weekend ER")
        ws.cell(row=1, column=start_col+5, value="Weekend ward")
        ws.cell(row=1, column=start_col+6, value="Total Weekend")
        for i, doctor in enumerate(doctors):
            row_num = i + 2
            ws.cell(row=row_num, column=start_col, value=doctor)
            col_map = {col: idx+2 for idx, col in enumerate(df.columns)}
            # Store formulas for each type for total columns
            weekday_formula = []
            weekend_formula = []
            # Weekday ER
            for j, (period, shift_type) in enumerate([
                    ("weekday", "ER"), ("weekday", "ward")]):
                count_formula = []
                for r, date in enumerate(dates):
                    is_wkend = is_weekend(date) or is_holiday(date)
                    if period == "weekday" and not is_wkend:
                        for col in df.columns:
                            if col.startswith(shift_type):
                                cell = ws.cell(
                                    row=r+2, column=col_map[col]).coordinate
                                count_formula.append(f'--({cell}="{doctor}")')
                if count_formula:
                    formula = f'=SUM({" ".join(count_formula)})'
                else:
                    formula = '=0'
                ws.cell(row=row_num, column=start_col+1+j, value=formula)
                weekday_formula.append(f'({formula[1:]})')
            # Total Weekday
            if weekday_formula:
                total_weekday_formula = f'=SUM({"+".join(weekday_formula)})'
            else:
                total_weekday_formula = '=0'
            ws.cell(row=row_num, column=start_col +
                    3, value=total_weekday_formula)
            # Weekend ER
            for j, (period, shift_type) in enumerate([
                    ("weekend", "ER"), ("weekend", "ward")]):
                count_formula = []
                for r, date in enumerate(dates):
                    is_wkend = is_weekend(date) or is_holiday(date)
                    if period == "weekend" and is_wkend:
                        for col in df.columns:
                            if col.startswith(shift_type):
                                cell = ws.cell(
                                    row=r+2, column=col_map[col]).coordinate
                                count_formula.append(f'--({cell}="{doctor}")')
                if count_formula:
                    formula = f'=SUM({" ".join(count_formula)})'
                else:
                    formula = '=0'
                ws.cell(row=row_num, column=start_col+4+j, value=formula)
                weekend_formula.append(f'({formula[1:]})')
            # Total Weekend
            if weekend_formula:
                total_weekend_formula = f'=SUM({"+".join(weekend_formula)})'
            else:
                total_weekend_formula = '=0'
            ws.cell(row=row_num, column=start_col +
                    6, value=total_weekend_formula)
        expected_row = len(doctors) + 3
        ws.cell(row=expected_row-1, column=start_col, value="Expected")
        ws.cell(row=expected_row, column=start_col, value="Doctor")
        ws.cell(row=expected_row, column=start_col+1, value="Weekday ER")
        ws.cell(row=expected_row, column=start_col+2, value="Weekday ward")
        ws.cell(row=expected_row, column=start_col+3, value="Total Weekday")
        ws.cell(row=expected_row, column=start_col+4, value="Weekend ER")
        ws.cell(row=expected_row, column=start_col+5, value="Weekend ward")
        ws.cell(row=expected_row, column=start_col+6, value="Total Weekend")
        for i, doctor in enumerate(doctors):
            ws.cell(row=expected_row+1+i, column=start_col, value=doctor)
            ws.cell(row=expected_row+1+i, column=start_col+1,
                    value=adjusted_doctor_data[doctor]["weekday"]["ER"])
            ws.cell(row=expected_row+1+i, column=start_col+2,
                    value=adjusted_doctor_data[doctor]["weekday"]["ward"])
            ws.cell(row=expected_row+1+i, column=start_col+3,
                    value=adjusted_doctor_data[doctor]["weekday"]["ER"] + adjusted_doctor_data[doctor]["weekday"]["ward"])
            ws.cell(row=expected_row+1+i, column=start_col+4,
                    value=adjusted_doctor_data[doctor]["weekend"]["ER"])
            ws.cell(row=expected_row+1+i, column=start_col+5,
                    value=adjusted_doctor_data[doctor]["weekend"]["ward"])
            ws.cell(row=expected_row+1+i, column=start_col+6,
                    value=adjusted_doctor_data[doctor]["weekend"]["ER"] + adjusted_doctor_data[doctor]["weekend"]["ward"])
        light_orange = PatternFill(
            start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")
        light_green = PatternFill(
            start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
        for idx, date in enumerate(dates):
            excel_row = idx + 2
            period = ws.cell(row=excel_row, column=2).value
            fill = light_orange if period == "Weekend" else light_green
            for col in range(1, len(df.columns) + 3):
                ws.cell(row=excel_row, column=col).fill = fill
                ws.cell(row=excel_row, column=col).border = None
    print(f"Schedule saved to {os.path.abspath(filename)}")
